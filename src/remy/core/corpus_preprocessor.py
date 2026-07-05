#!/usr/bin/env python
"""Normalize mixed source files into a clean UTF-8 text corpus for AHS learning.

This is a transport/canonicalization stage, not a truth or semantic stage.
It converts file formats into plain text and writes a manifest that preserves
source provenance. It does not grant answer authority.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
from html.parser import HTMLParser
import importlib
import json
import re
import time
import zipfile
from pathlib import Path
from xml.etree import ElementTree


SCHEMA = "aura_ahs_unified_corpus_preprocessor_v0"
ANCHOR_SCHEMA = "aura_ahs_clean_corpus_anchor_v0"
DEFAULT_EXTENSIONS = (
    ".txt",
    ".md",
    ".markdown",
    ".rst",
    ".html",
    ".htm",
    ".xml",
    ".json",
    ".jsonl",
    ".csv",
    ".docx",
    ".pdf",
    ".xlsx",
    ".xml-p1p194007",
)


class ReadableMarkupParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.skip_stack: list[str] = []
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag in {"script", "style", "nav", "header", "footer", "aside", "noscript"}:
            self.skip_stack.append(tag)
            return
        if tag == "li":
            self.parts.append("\n- ")
            return
        if re.fullmatch(r"h[1-6]", tag):
            level = int(tag[1])
            self.parts.append("\n" + ("#" * min(level, 6)) + " ")
            return
        if tag in {"td", "th"}:
            self.parts.append(" | ")
            return
        if tag in {"p", "br", "tr", "div", "section", "article"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if self.skip_stack and self.skip_stack[-1] == tag:
            self.skip_stack.pop()
            return
        if tag in {"p", "li", "tr", "div", "section", "article"} or re.fullmatch(r"h[1-6]", tag):
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self.skip_stack:
            return
        text = " ".join(str(data or "").split())
        if text:
            self.parts.append(text)
            self.parts.append(" ")

    def text(self) -> str:
        return normalize_text("".join(self.parts))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", required=True, action="append")
    parser.add_argument("--output-root", default="target/aura-local/ahs-v0/clean-corpus")
    parser.add_argument("--manifest-output", default="")
    parser.add_argument("--report-output", default="")
    parser.add_argument("--progress-output", default="")
    parser.add_argument("--file-extensions", default=",".join(ext.lstrip(".") for ext in DEFAULT_EXTENSIONS))
    parser.add_argument("--max-files", type=int, default=0)
    parser.add_argument("--max-bytes-per-file", type=int, default=0)
    parser.add_argument("--min-clean-chars", type=int, default=64)
    parser.add_argument("--source-origin-channel", default="local_file")
    parser.add_argument("--source-acquisition-surface", default="manual_upload")
    parser.add_argument("--verification-mode-hint", default="conflict_scan")
    parser.add_argument(
        "--extractor-provider",
        choices=("auto", "docling", "markitdown", "gensim_wiki", "built_in"),
        default="built_in",
    )
    parser.add_argument("--clear-output", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    readout = build_clean_corpus(
        source_roots=[Path(item) for item in args.source_root],
        output_root=Path(args.output_root),
        manifest_output=Path(args.manifest_output) if args.manifest_output else None,
        report_output=Path(args.report_output) if args.report_output else None,
        progress_output=Path(args.progress_output) if args.progress_output else None,
        extensions=parse_extensions(args.file_extensions),
        max_files=max(0, args.max_files),
        max_bytes_per_file=max(0, args.max_bytes_per_file),
        min_clean_chars=max(0, args.min_clean_chars),
        source_origin_channel=safe_lineage_token(args.source_origin_channel, "local_file"),
        source_acquisition_surface=safe_lineage_token(args.source_acquisition_surface, "manual_upload"),
        verification_mode_hint=safe_lineage_token(args.verification_mode_hint, "conflict_scan"),
        extractor_provider=args.extractor_provider,
        clear_output=bool(args.clear_output),
    )
    print(json.dumps(readout, ensure_ascii=False, indent=2, sort_keys=True))


def build_clean_corpus(
    *,
    source_root: Path | None = None,
    source_roots: list[Path] | None = None,
    output_root: Path,
    manifest_output: Path | None,
    report_output: Path | None,
    progress_output: Path | None,
    extensions: set[str],
    max_files: int,
    max_bytes_per_file: int,
    min_clean_chars: int,
    source_origin_channel: str,
    source_acquisition_surface: str,
    verification_mode_hint: str,
    extractor_provider: str,
    clear_output: bool,
    progress_callback=None,
) -> dict:
    roots = [Path(item) for item in (source_roots or ([source_root] if source_root is not None else []))]
    if not roots:
        raise ValueError("at least one source root is required")
    source_root_label = ";".join(str(item) for item in roots)
    if clear_output and output_root.exists():
        for path in sorted(output_root.rglob("*"), reverse=True):
            if path.is_file():
                path.unlink()
            elif path.is_dir():
                path.rmdir()
    output_root.mkdir(parents=True, exist_ok=True)
    manifest_output = manifest_output or output_root / "manifest.jsonl"
    report_output = report_output or output_root / "report.json"
    manifest_output.parent.mkdir(parents=True, exist_ok=True)
    report_output.parent.mkdir(parents=True, exist_ok=True)

    files = collect_files_from_roots(roots, extensions)
    total_files_found = len(files)
    truncated_count = 0
    if max_files and total_files_found > max_files:
        truncated_count = total_files_found - max_files
    selected = files[:max_files] if max_files and len(files) > max_files else files
    selected_source_bytes = sum(safe_size(path) for _root, path in selected)
    records = []
    source_bytes = 0
    clean_bytes = 0
    skipped = []
    _all_skipped = []  # store ALL skipped, not capped at 64

    def emit_clean_progress(**kwargs):
        payload = write_clean_progress(progress_output, **kwargs)
        if progress_callback is not None:
            progress_callback(payload)

    emit_clean_progress(
        source_root=source_root_label,
        output_root=output_root,
        status="running",
        processed_file_count=0,
        selected_file_count=len(selected),
        processed_bytes=0,
        selected_source_bytes=selected_source_bytes,
        clean_bytes_written=0,
        skipped_file_count=0,
        current_source_path="",
        message=f"Clean corpus conversion started. {truncated_count} files truncated, {total_files_found} total found.",
    )
    # Preflight: check external extractor availability before processing any files
    _preflight_extractors(extractor_provider)
    for index, item in enumerate(selected, start=1):
        root, path = item
        raw_size = safe_size(path)
        completed_source_bytes = source_bytes
        source_bytes += raw_size
        rel = safe_relative(path, root)
        if not rel.name:
            rel = Path(path.name)
        digest = sha256_hex(str(path).encode("utf-8") + str(raw_size).encode("ascii"))
        out_path = output_root / rel.with_suffix(rel.suffix + f".{digest[:12]}.txt")
        out_path.parent.mkdir(parents=True, exist_ok=True)
        clean_text = ""
        clean_size = 0
        streamed = False
        emit_clean_progress(
            source_root=source_root_label,
            output_root=output_root,
            status="running",
            processed_file_count=index - 1,
            selected_file_count=len(selected),
            processed_bytes=completed_source_bytes,
            selected_source_bytes=selected_source_bytes,
            clean_bytes_written=clean_bytes,
            skipped_file_count=len(skipped),
            current_source_path=str(path),
            message="Converting source file to clean UTF-8 text.",
        )
        try:
            if is_xml_like(path) and max_bytes_per_file == 0 and raw_size > 256 * 1024 * 1024:
                # Try built-in byte-level parser first (5-10x faster than gensim).
                # Gensim article-by-article DOM parsing is CPU-bound and slow for
                # 200K+ article dumps. Fall back to gensim only if built-in fails.
                stream_size, warnings = stream_mediawiki_text_payloads_to_file(
                    path,
                    out_path,
                    min_clean_chars,
                    progress_callback=lambda file_bytes_read, file_clean_bytes: emit_clean_progress(
                        source_root=source_root_label,
                        output_root=output_root,
                        status="running",
                        processed_file_count=index - 1,
                        selected_file_count=len(selected),
                        processed_bytes=completed_source_bytes + file_bytes_read,
                        selected_source_bytes=selected_source_bytes,
                        clean_bytes_written=clean_bytes + file_clean_bytes,
                        skipped_file_count=len(skipped),
                        current_source_path=str(path),
                        message="Streaming large XML text payloads into clean UTF-8 corpus (built-in parser).",
                    ),
                )
                if stream_size <= 0:
                    stream_size, warnings = stream_external_mediawiki_text_to_file(
                        path,
                        out_path,
                        min_clean_chars,
                        extractor_provider=extractor_provider,
                        progress_callback=lambda file_bytes_read, file_clean_bytes: emit_clean_progress(
                            source_root=source_root_label,
                            output_root=output_root,
                            status="running",
                            processed_file_count=index - 1,
                            selected_file_count=len(selected),
                            processed_bytes=completed_source_bytes + file_bytes_read,
                            selected_source_bytes=selected_source_bytes,
                            clean_bytes_written=clean_bytes + file_clean_bytes,
                            skipped_file_count=len(skipped),
                            current_source_path=str(path),
                            message="Streaming large XML text payloads into clean UTF-8 corpus (gensim fallback).",
                        ),
                    )
                if stream_size >= min_clean_chars:
                    clean_size = stream_size
                    streamed = True
                else:
                    clean_text, warnings = extract_clean_text(
                        path,
                        max_bytes_per_file=max_bytes_per_file,
                        extractor_provider=extractor_provider,
                    )
            else:
                clean_text, warnings = extract_clean_text(
                    path,
                    max_bytes_per_file=max_bytes_per_file,
                    extractor_provider=extractor_provider,
                )
        except Exception as exc:  # noqa: BLE001 - report, do not fail whole corpus
            error_entry = {"path": str(path), "reason": "extract_failed", "error": str(exc)}
            skipped.append(error_entry)
            _all_skipped.append(error_entry)
            emit_clean_progress(
                source_root=source_root_label,
                output_root=output_root,
                status="running",
                processed_file_count=index,
                selected_file_count=len(selected),
                processed_bytes=source_bytes,
                selected_source_bytes=selected_source_bytes,
                clean_bytes_written=clean_bytes,
                skipped_file_count=len(skipped),
                current_source_path=str(path),
                message="Source file skipped after extraction failure.",
            )
            continue
        if not streamed:
            clean_text = normalize_text(clean_text)
            if len(clean_text) < min_clean_chars:
                skipped.append({"path": str(path), "reason": "clean_text_too_short", "clean_chars": len(clean_text)})
                emit_clean_progress(
                    source_root=source_root_label,
                    output_root=output_root,
                    status="running",
                    processed_file_count=index,
                    selected_file_count=len(selected),
                    processed_bytes=source_bytes,
                    selected_source_bytes=selected_source_bytes,
                    clean_bytes_written=clean_bytes,
                    skipped_file_count=len(skipped),
                    current_source_path=str(path),
                    message="Source file skipped because clean text was too short.",
                )
                continue
            with out_path.open("w", encoding="utf-8", errors="strict", newline="\n") as output:
                output.write(clean_text)
                output.write("\n")
            write_text_surface_anchors(
                source_path=path,
                clean_path=out_path,
                clean_text=clean_text,
            )
            clean_size = len(clean_text.encode("utf-8"))
        clean_bytes += clean_size
        quality_text = clean_text
        if streamed:
            quality_text = _read_quality_window(out_path)
        clean_surface_quality = assess_clean_surface_quality(
            quality_text,
            source_size_bytes=raw_size,
            clean_size_bytes=clean_size,
            warnings=warnings,
        )
        emit_clean_progress(
            source_root=source_root_label,
            output_root=output_root,
            status="running",
            processed_file_count=index,
            selected_file_count=len(selected),
            processed_bytes=source_bytes,
            selected_source_bytes=selected_source_bytes,
            clean_bytes_written=clean_bytes,
            skipped_file_count=len(skipped),
            current_source_path=str(path),
            message="Source file converted into clean UTF-8 text.",
        )
        records.append(
            {
                "schema": "aura_ahs_clean_corpus_record_v0",
                "record_index": index,
                "source_path": str(path),
                "source_size_bytes": raw_size,
                "clean_path": str(out_path),
                "clean_size_bytes": clean_size,
                "source_sha256": file_sha256(path, max_bytes=max_bytes_per_file),
                "clean_sha256": sha256_hex(clean_text.encode("utf-8")),
                "clean_stream_sha256": file_sha256(out_path, max_bytes=0) if streamed else "",
                "lineage_contract_schema": "aura_source_lineage_contract_v0",
                "source_origin_channel": source_origin_channel,
                "source_acquisition_surface": source_acquisition_surface,
                "source_origin_event_id_u64": stable_u64(
                    f"{source_origin_channel}:{source_acquisition_surface}:{path}:{raw_size}"
                ),
                "source_origin_record_hash_u64": stable_u64(
                    f"{source_origin_channel}:{path}:{raw_size}:{clean_size}"
                ),
                "verification_mode_hint": verification_mode_hint,
                "extractor_provider_requested": extractor_provider,
                "warnings": warnings,
                "clean_surface_quality": clean_surface_quality,
                "candidate_field_only": True,
                "answer_permission_granted": False,
                "truth_asserted": False,
            }
        )
    with manifest_output.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True))
            handle.write("\n")
    readout = {
        "schema": SCHEMA,
        "source_root": str(roots[0]) if len(roots) == 1 else source_root_label,
        "source_roots": [str(item) for item in roots],
        "output_root": str(output_root),
        "manifest_output": str(manifest_output),
        "report_output": str(report_output),
        "source_file_count": len(files),
        "total_files_found": total_files_found,
        "selected_file_count": len(selected),
        "truncated_due_to_max_files": truncated_count,
        "clean_file_count": len(records),
        "skipped_file_count": len(skipped),
        "source_bytes_selected": source_bytes,
        "clean_bytes_written": clean_bytes,
        "extensions": sorted(extensions),
        "min_clean_chars": min_clean_chars,
        "max_bytes_per_file": max_bytes_per_file,
        "lineage_contract_schema": "aura_source_lineage_contract_v0",
        "source_origin_channel": source_origin_channel,
        "source_acquisition_surface": source_acquisition_surface,
        "verification_mode_hint": verification_mode_hint,
        "extractor_provider_requested": extractor_provider,
        "clean_surface_quality": summarize_clean_surface_quality(records),
        "skipped": _all_skipped,
        "skipped_count": len(_all_skipped),
        "skipped_surface": _all_skipped[:64],
        "candidate_field_only": True,
        "answer_permission_granted": False,
        "truth_asserted": False,
    }
    report_output.write_text(json.dumps(readout, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    emit_clean_progress(
        source_root=source_root_label,
        output_root=output_root,
        status="succeeded",
        processed_file_count=len(selected),
        selected_file_count=len(selected),
        processed_bytes=source_bytes,
        selected_source_bytes=selected_source_bytes,
        clean_bytes_written=clean_bytes,
        skipped_file_count=len(skipped),
        current_source_path="",
        message="Clean corpus conversion finished.",
    )
    return readout


def write_clean_progress(
    progress_output: Path | None,
    *,
    source_root: Path | str,
    output_root: Path,
    status: str,
    processed_file_count: int,
    selected_file_count: int,
    processed_bytes: int,
    selected_source_bytes: int,
    clean_bytes_written: int,
    skipped_file_count: int,
    current_source_path: str,
    message: str,
) -> dict:
    if not progress_output:
        return {}
    progress_output.parent.mkdir(parents=True, exist_ok=True)
    progress_percent = 100.0 if selected_source_bytes <= 0 and status == "succeeded" else (
        min(100.0, max(0.0, (float(processed_bytes) / float(selected_source_bytes)) * 100.0))
        if selected_source_bytes > 0
        else 0.0
    )
    payload = {
        "schema": "aura_ahs_clean_corpus_progress_v0",
        "source_root": str(source_root),
        "output_root": str(output_root),
        "status": status,
        "processed_file_count": processed_file_count,
        "selected_file_count": selected_file_count,
        "processed_bytes": processed_bytes,
        "selected_source_bytes": selected_source_bytes,
        "clean_bytes_written": clean_bytes_written,
        "skipped_file_count": skipped_file_count,
        "current_source_path": current_source_path,
        "progress_percent": progress_percent,
        "message": message,
        "candidate_field_only": True,
        "answer_permission_granted": False,
        "truth_asserted": False,
    }
    write_json_atomic(progress_output, payload)
    return payload


def write_json_atomic(path: Path, payload: dict) -> None:
    tmp_path = path.with_name(f"{path.name}.tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    for _ in range(8):
        try:
            tmp_path.replace(path)
            return
        except PermissionError:
            time.sleep(0.05)
    try:
        tmp_path.unlink()
    except OSError:
        pass


def _read_quality_window(path: Path, *, limit: int = 65536) -> str:
    try:
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            return handle.read(limit)
    except OSError:
        return ""


def assess_clean_surface_quality(
    text: str,
    *,
    source_size_bytes: int,
    clean_size_bytes: int,
    warnings: list[str],
) -> dict:
    content = str(text or "")
    chars = len(content)
    nonspace = sum(1 for ch in content if not ch.isspace())
    letters_digits = sum(1 for ch in content if ch.isalpha() or ch.isdigit())
    replacement_count = content.count("\ufffd")
    control_count = sum(1 for ch in content if ord(ch) < 32 and ch not in "\n\t\r")
    lines = [line.strip() for line in content.splitlines() if line.strip()]
    structure_count = sum(
        1
        for line in lines
        if line.startswith(("#", "- ", "|", "---")) or " | " in line
    )
    html_noise_count = len(re.findall(r"</?[a-zA-Z][^>]*>|&[a-zA-Z]+;", content))
    readable_ratio = (letters_digits / nonspace) if nonspace else 0.0
    replacement_ratio = (replacement_count / chars) if chars else 1.0
    clean_to_source_ratio = (clean_size_bytes / source_size_bytes) if source_size_bytes > 0 else 0.0

    score = 100
    reasons: list[str] = []
    if clean_size_bytes <= 0 or chars <= 0:
        score -= 100
        reasons.append("EmptyCleanSurface")
    if chars < 64:
        score -= 25
        reasons.append("VeryShortCleanSurface")
    if readable_ratio < 0.55:
        score -= 20
        reasons.append("LowReadableCharacterRatio")
    if replacement_ratio > 0.005:
        score -= 25
        reasons.append("UnicodeReplacementCharacters")
    if control_count:
        score -= 20
        reasons.append("ControlCharactersPresent")
    if html_noise_count:
        score -= min(30, html_noise_count * 5)
        reasons.append("MarkupNoisePresent")
    if source_size_bytes >= 4096 and clean_to_source_ratio < 0.01:
        score -= 20
        reasons.append("VeryLowCleanToSourceRatio")
    if any("fallback" in str(item).lower() or "failed" in str(item).lower() for item in warnings or []):
        score -= 10
        reasons.append("ExtractorFallbackOrFailureWarning")
    if structure_count == 0 and any(
        marker in (warnings or [])
        for marker in (
            "external_extractor:docling",
            "external_extractor:markitdown",
            "canonical_markdown_surface",
        )
    ):
        score -= 5
        reasons.append("NoStructureMarkers")

    score = max(0, min(100, score))
    if score >= 80:
        decision = "Ready"
    elif score >= 55:
        decision = "Review"
    else:
        decision = "Blocked"
    return {
        "schema": "aura_clean_surface_quality_v0",
        "decision": decision,
        "score": score,
        "reasons": reasons,
        "char_count": chars,
        "line_count": len(lines),
        "structure_marker_count": structure_count,
        "readable_character_ratio": round(readable_ratio, 4),
        "replacement_character_count": replacement_count,
        "control_character_count": control_count,
        "markup_noise_count": html_noise_count,
        "clean_to_source_ratio": round(clean_to_source_ratio, 6),
        "candidate_only": True,
        "answer_permission_granted": False,
        "truth_asserted": False,
    }


def summarize_clean_surface_quality(records: list[dict]) -> dict:
    qualities = [
        row.get("clean_surface_quality")
        for row in records
        if isinstance(row.get("clean_surface_quality"), dict)
    ]
    if not qualities:
        return {
            "schema": "aura_clean_surface_quality_summary_v0",
            "decision": "Blocked",
            "sample_count": 0,
            "average_score": 0,
            "minimum_score": 0,
            "ready_count": 0,
            "review_count": 0,
            "blocked_count": 0,
            "answer_permission_granted": False,
            "truth_asserted": False,
        }
    scores = [int(item.get("score") or 0) for item in qualities]
    ready_count = sum(1 for item in qualities if item.get("decision") == "Ready")
    review_count = sum(1 for item in qualities if item.get("decision") == "Review")
    blocked_count = sum(1 for item in qualities if item.get("decision") == "Blocked")
    decision = "Ready" if blocked_count == 0 and review_count == 0 else ("Review" if blocked_count == 0 else "Blocked")
    reasons = sorted({reason for item in qualities for reason in (item.get("reasons") or [])})
    return {
        "schema": "aura_clean_surface_quality_summary_v0",
        "decision": decision,
        "sample_count": len(qualities),
        "average_score": round(sum(scores) / len(scores), 2),
        "minimum_score": min(scores),
        "ready_count": ready_count,
        "review_count": review_count,
        "blocked_count": blocked_count,
        "reasons": reasons,
        "candidate_only": True,
        "answer_permission_granted": False,
        "truth_asserted": False,
    }


def collect_files(root: Path, extensions: set[str]) -> list[Path]:
    if root.is_file():
        return [root] if suffix_allowed(root, extensions) else []
    files = [path for path in root.rglob("*") if path.is_file() and suffix_allowed(path, extensions)]
    files.sort()
    return files


def collect_files_from_roots(roots: list[Path], extensions: set[str]) -> list[tuple[Path, Path]]:
    seen: set[Path] = set()
    files: list[tuple[Path, Path]] = []
    for root in roots:
        for path in collect_files(root, extensions):
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            files.append((root, path))
    files.sort(key=lambda item: str(item[1]).lower())
    return files


def suffix_allowed(path: Path, extensions: set[str]) -> bool:
    lower = path.name.lower()
    return any(lower.endswith(ext) for ext in extensions)


def safe_lineage_token(value: str, fallback: str) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9_.:-]+", "_", text).strip("_.:-")
    return text or fallback


def stable_u64(value: str) -> int:
    return int.from_bytes(hashlib.sha256(str(value).encode("utf-8")).digest()[:8], "big")


def is_xml_like(path: Path) -> bool:
    lower = path.name.lower()
    return lower.endswith(".xml") or ".xml-" in lower


def _preflight_extractors(provider):
    """Check external extractor availability before processing. Fails fast."""
    if provider in ("built_in", "auto"):
        return
    failures = []
    for item in (["docling", "markitdown"] if provider == "auto" else [provider]):
        try:
            if item == "docling":
                importlib.import_module("docling.document_converter")
            elif item == "markitdown":
                importlib.import_module("markitdown")
        except ModuleNotFoundError as exc:
            failures.append(f"{item}: {exc}")
    if failures and provider != "auto":
        raise ModuleNotFoundError(f"Extractor(s) unavailable: {'; '.join(failures)}")
    if failures:
        print(f"Warning: Some extractors unavailable, will use fallback: {'; '.join(failures)}")


def extract_clean_text(
    path: Path,
    *,
    max_bytes_per_file: int,
    extractor_provider: str = "auto",
) -> tuple[str, list[str]]:
    lower = path.name.lower()
    external = external_extractor_text(
        path,
        provider=extractor_provider,
        max_bytes_per_file=max_bytes_per_file,
    )
    if external is not None:
        text, warnings = external
        return normalize_text(text), warnings
    if lower.endswith(".docx"):
        return docx_to_text(path)
    if lower.endswith(".pdf"):
        return pdf_to_text(path, max_bytes_per_file=max_bytes_per_file)
    if lower.endswith(".html") or lower.endswith(".htm"):
        return markup_to_text(read_text(path, max_bytes_per_file=max_bytes_per_file)), ["html_markup_stripped"]
    if lower.endswith(".xml") or ".xml-" in lower:
        return xml_to_text(
            path,
            max_bytes_per_file=max_bytes_per_file,
            extractor_provider=extractor_provider,
        )
    if lower.endswith(".json"):
        return json_to_text(path, max_bytes_per_file=max_bytes_per_file)
    if lower.endswith(".jsonl"):
        return jsonl_to_text(path, max_bytes_per_file=max_bytes_per_file)
    if lower.endswith(".csv"):
        return csv_to_text(path, max_bytes_per_file=max_bytes_per_file)
    if lower.endswith(".xlsx"):
        return xlsx_to_text(path)
    return read_text(path, max_bytes_per_file=max_bytes_per_file), ["plain_text_normalized"]


def external_extractor_text(
    path: Path,
    *,
    provider: str,
    max_bytes_per_file: int,
) -> tuple[str, list[str]] | None:
    lower = path.name.lower()
    if provider == "built_in":
        return None
    if max_bytes_per_file > 0:
        return None
    if not lower.endswith((".pdf", ".docx", ".pptx", ".xlsx", ".html", ".htm")):
        return None
    providers = [provider] if provider in {"docling", "markitdown"} else ["docling", "markitdown"]
    failures = []
    for item in providers:
        try:
            if item == "docling":
                text = docling_to_text(path)
            elif item == "markitdown":
                text = markitdown_to_text(path)
            else:
                continue
        except Exception as exc:  # noqa: BLE001 - external optional adapter
            failures.append(f"{item}_unavailable:{type(exc).__name__}")
            continue
        clean = normalize_text(text)
        if clean:
            warnings = [f"external_extractor:{item}"]
            warnings.extend(failures)
            return clean, warnings
    return None


def docling_to_text(path: Path) -> str:
    module = importlib.import_module("docling.document_converter")
    converter_class = getattr(module, "DocumentConverter")
    result = converter_class().convert(str(path))
    document = getattr(result, "document", None)
    if document is None:
        return str(result)
    if hasattr(document, "export_to_markdown"):
        return str(document.export_to_markdown())
    if hasattr(document, "export_to_text"):
        return str(document.export_to_text())
    return str(document)


def markitdown_to_text(path: Path) -> str:
    module = importlib.import_module("markitdown")
    converter_class = getattr(module, "MarkItDown")
    result = converter_class().convert(str(path))
    text = getattr(result, "text_content", None)
    if text is not None:
        return str(text)
    return str(result)


def docx_to_text(path: Path) -> tuple[str, list[str]]:
    warnings = ["docx_zip_xml_extracted"]
    with zipfile.ZipFile(path) as archive:
        names = [name for name in archive.namelist() if name.startswith("word/") and name.endswith(".xml")]
        preferred = [name for name in names if name == "word/document.xml"] or names
        parts = []
        for name in preferred[:8]:
            blob = archive.read(name)
            try:
                root = ElementTree.fromstring(blob)
                for node in root.iter():
                    if node.text:
                        parts.append(node.text)
            except ElementTree.ParseError:
                parts.append(re.sub(r"<[^>]+>", " ", blob.decode("utf-8", errors="replace")))
        return normalize_text("\n".join(parts)), warnings


def pdf_to_text(path: Path, *, max_bytes_per_file: int) -> tuple[str, list[str]]:
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        return normalize_text("\n".join(parts)), ["pdf_extracted_with_pypdf"]
    except Exception:
        raw = read_bytes(path, max_bytes=max_bytes_per_file)
        printable = re.findall(rb"[\w\s.,;:!?()\-/]{8,}", raw)
        text = b"\n".join(printable).decode("utf-8", errors="replace")
        return normalize_text(text), ["pdf_lightweight_fallback_used"]


def json_to_text(path: Path, *, max_bytes_per_file: int) -> tuple[str, list[str]]:
    try:
        payload = json.loads(read_text(path, max_bytes_per_file=max_bytes_per_file))
    except json.JSONDecodeError:
        return read_text(path, max_bytes_per_file=max_bytes_per_file), ["json_parse_failed_plain_text_fallback"]
    return normalize_text(json_payload_to_markdown(payload)), [
        "json_scalar_values_extracted",
        "canonical_markdown_surface",
    ]


def jsonl_to_text(path: Path, *, max_bytes_per_file: int) -> tuple[str, list[str]]:
    values: list[str] = []
    failures = 0
    for line in read_text(path, max_bytes_per_file=max_bytes_per_file).splitlines():
        text = line.strip()
        if not text:
            continue
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            failures += 1
            values.append(text)
            continue
        rendered = json_payload_to_markdown(payload)
        if rendered:
            values.append(rendered)
    warnings = ["jsonl_scalar_values_extracted", "canonical_markdown_surface"]
    if failures:
        warnings.append(f"jsonl_line_parse_failures:{failures}")
    return normalize_text("\n".join(values)), warnings


def iter_json_scalar_values(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        return [text] if text else []
    if isinstance(value, bool):
        return []
    if isinstance(value, (int, float)):
        return [str(value)]
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            result.extend(iter_json_scalar_values(item))
        return result
    if isinstance(value, dict):
        result = []
        for item in value.values():
            result.extend(iter_json_scalar_values(item))
        return result
    return []


def json_payload_to_markdown(value) -> str:
    rows: list[str] = []

    def emit(item, path: str = "") -> None:
        if item is None or isinstance(item, bool):
            return
        if isinstance(item, str):
            text = item.strip()
            if not text:
                return
            if path.lower() in {"title", "name", "heading"}:
                rows.append(f"# {text}")
            else:
                rows.append(f"- {text}" if path else text)
            return
        if isinstance(item, (int, float)):
            rows.append(f"- {item}" if path else str(item))
            return
        if isinstance(item, list):
            if item and all(isinstance(row, dict) for row in item):
                table = dict_rows_to_markdown_table(item)
                if table:
                    rows.append(table)
                    return
            for child in item:
                emit(child, path)
            return
        if isinstance(item, dict):
            for key, child in item.items():
                emit(child, str(key).strip())

    emit(value)
    return "\n".join(rows)


def dict_rows_to_markdown_table(items) -> str:
    keys: list[str] = []
    for item in items:
        if not isinstance(item, dict):
            return ""
        for key in item.keys():
            text = str(key).strip()
            if text and text not in keys:
                keys.append(text)
    if not keys:
        return ""
    rows = [keys]
    for item in items:
        rows.append([str(item.get(key) or "").strip() for key in keys])
    return rows_to_markdown_table(rows)


def rows_to_markdown_table(rows: list[list[str]]) -> str:
    normalized = [[str(cell or "").strip() for cell in row] for row in rows if row]
    if len(normalized) < 2:
        return ""
    width = max(len(row) for row in normalized)
    if width < 2:
        return ""
    padded = [row + [""] * (width - len(row)) for row in normalized]
    header = padded[0]
    body = padded[1:]
    if not any(any(cell for cell in row) for row in body):
        return ""
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in range(width)) + " |",
    ]
    lines.extend("| " + " | ".join(row) + " |" for row in body)
    return "\n".join(lines)


def rows_to_canonical_markdown(rows: list[list[str]]) -> str:
    normalized = [[str(cell or "").strip() for cell in row if str(cell or "").strip()] for row in rows]
    normalized = [row for row in normalized if row]
    if not normalized:
        return ""
    parts: list[str] = []
    index = 0
    while index < len(normalized):
        width = len(normalized[index])
        run_end = index + 1
        while run_end < len(normalized) and len(normalized[run_end]) == width:
            run_end += 1
        run = normalized[index:run_end]
        if width >= 2 and len(run) >= 2:
            table = rows_to_markdown_table(run)
            if table:
                parts.append(table)
            else:
                parts.extend("- " + " | ".join(row) for row in run)
        else:
            for row in run:
                if len(row) == 2:
                    parts.append(f"- {row[0]}: {row[1]}")
                else:
                    parts.append("- " + " | ".join(row))
        index = run_end
    return "\n\n".join(parts)


def csv_to_text(path: Path, *, max_bytes_per_file: int) -> tuple[str, list[str]]:
    text = read_text(path, max_bytes_per_file=max_bytes_per_file)
    reader = csv.reader(io.StringIO(text))
    rows = []
    for row in reader:
        cells = [cell.strip() for cell in row if cell and cell.strip()]
        if cells:
            rows.append(cells)
    rendered = rows_to_canonical_markdown(rows)
    return normalize_text(rendered), ["csv_cell_values_extracted", "canonical_markdown_surface"]


def xlsx_to_text(path: Path) -> tuple[str, list[str]]:
    """Extract text from Excel .xlsx using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        return read_text(path, max_bytes_per_file=0), ["xlsx_openpyxl_unavailable"]
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    sheet_rows.append(cells)
            if sheet_rows:
                rendered = rows_to_canonical_markdown(sheet_rows)
                parts.append(f"## {sheet_name}\n{rendered}")
        wb.close()
        if not parts:
            return "", ["xlsx_no_data"]
        return normalize_text("\n\n".join(parts)), ["xlsx_openpyxl_extracted", "canonical_markdown_surface"]
    except Exception:
        return read_text(path, max_bytes_per_file=0), ["xlsx_read_fallback"]


def xml_to_text(
    path: Path,
    *,
    max_bytes_per_file: int,
    extractor_provider: str = "auto",
) -> tuple[str, list[str]]:
    medline_text = medlineplus_xml_to_text(path, max_bytes_per_file=max_bytes_per_file)
    if medline_text:
        return medline_text, ["medlineplus_payload_extracted"]
    external = external_mediawiki_text(path, extractor_provider=extractor_provider)
    if external is not None:
        return external
    text_payloads = extract_mediawiki_text_payloads(path, max_bytes_per_file=max_bytes_per_file)
    if text_payloads:
        clean_payloads = [strip_wiki_markup(payload) for payload in text_payloads]
        return normalize_text("\n\n".join(clean_payloads)), ["mediawiki_text_payload_extracted", "wiki_markup_stripped"]
    generic = generic_xml_to_markdown(path, max_bytes_per_file=max_bytes_per_file)
    if generic:
        return normalize_text(generic), ["xml_markup_stripped", "canonical_markdown_surface"]
    return markup_to_text(read_text(path, max_bytes_per_file=max_bytes_per_file)), ["xml_markup_stripped"]


def generic_xml_to_markdown(path: Path, *, max_bytes_per_file: int) -> str:
    raw = read_bytes(path, max_bytes=max_bytes_per_file)
    try:
        root = ElementTree.fromstring(raw)
    except ElementTree.ParseError:
        return ""
    rows: list[str] = []
    for node in root.iter():
        text = " ".join(str(node.text or "").split()).strip()
        if not text:
            continue
        tag = local_name(node.tag).lower()
        if tag in {"title", "name", "heading", "h1"}:
            rows.append(f"# {text}")
        elif tag in {"subtitle", "h2", "section"}:
            rows.append(f"## {text}")
        else:
            rows.append(f"- {text}")
    return "\n".join(rows)


def medlineplus_xml_to_text(path: Path, *, max_bytes_per_file: int) -> str:
    raw = read_bytes(path, max_bytes=max_bytes_per_file)
    if b"<health-topics" not in raw[:8192] and b"<health-topic" not in raw[:8192]:
        return ""
    try:
        root = ElementTree.fromstring(raw)
    except ElementTree.ParseError:
        return ""
    if local_name(root.tag) != "health-topics":
        return ""
    topic_blocks = []
    for topic in root:
        if local_name(topic.tag) != "health-topic":
            continue
        block = medlineplus_topic_to_text(topic)
        if block:
            topic_blocks.append(block)
    return normalize_text("\n\n".join(topic_blocks))


def medlineplus_topic_to_text(topic: ElementTree.Element) -> str:
    parts: list[str] = []
    title = str(topic.attrib.get("title") or "").strip()
    meta_desc = str(topic.attrib.get("meta-desc") or "").strip()
    if title:
        parts.append(title)
    aliases = []
    for child in topic:
        if local_name(child.tag) == "also-called" and child.text and child.text.strip():
            aliases.append(child.text.strip())
    if aliases:
        parts.append("Also called: " + "; ".join(aliases))
    if meta_desc:
        parts.append(meta_desc)
    for child in topic:
        if local_name(child.tag) == "full-summary" and child.text and child.text.strip():
            parts.append(markup_to_text(child.text))
    return normalize_text("\n\n".join(parts))


def external_mediawiki_text(
    path: Path,
    *,
    extractor_provider: str,
) -> tuple[str, list[str]] | None:
    if extractor_provider == "built_in":
        return None
    if extractor_provider not in {"auto", "gensim_wiki"}:
        return None
    if not is_mediawiki_xml_source(path):
        return None
    try:
        articles = list(iter_gensim_mediawiki_articles(path, min_clean_chars=0, progress_callback=None))
    except Exception:
        return None
    if not articles:
        return None
    text = normalize_text("\n\n".join(articles))
    if not text:
        return None
    return text, [
        "external_extractor:gensim_segment_wiki",
        f"mediawiki_gensim_articles_extracted:{len(articles)}",
    ]


def stream_external_mediawiki_text_to_file(
    path: Path,
    out_path: Path,
    min_clean_chars: int,
    *,
    extractor_provider: str,
    progress_callback=None,
) -> tuple[int, list[str]]:
    if extractor_provider == "built_in":
        return 0, ["external_mediawiki_provider_disabled"]
    if extractor_provider not in {"auto", "gensim_wiki"}:
        return 0, [f"external_mediawiki_provider_not_selected:{extractor_provider}"]
    if not is_mediawiki_xml_source(path):
        return 0, ["external_mediawiki_not_detected"]
    total_bytes = 0
    article_count = 0
    out_path.parent.mkdir(parents=True, exist_ok=True)
    anchors_path = clean_anchor_path(out_path)
    try:
        with out_path.open("w", encoding="utf-8", errors="strict", newline="\n") as output, anchors_path.open(
            "w", encoding="utf-8", newline="\n"
        ) as anchors:
            for clean, article_anchors in iter_gensim_mediawiki_articles_with_anchors(
                path,
                min_clean_chars=min_clean_chars,
                progress_callback=lambda file_bytes_read, file_clean_bytes: (
                    progress_callback(file_bytes_read, total_bytes + file_clean_bytes)
                    if progress_callback is not None
                    else None
                ),
            ):
                write_article_surface_anchors(
                    anchors,
                    source_path=path,
                    clean_path=out_path,
                    article_text=clean,
                    article_start_offset=total_bytes,
                    article_anchors=article_anchors,
                )
                encoded_len = len(clean.encode("utf-8"))
                output.write(clean)
                output.write("\n\n")
                total_bytes += encoded_len + 2
                article_count += 1
    except Exception as exc:  # noqa: BLE001 - optional external adapter
        if article_count > 0:
            if progress_callback is not None:
                progress_callback(safe_size(path), total_bytes)
            warnings = [
                "external_extractor:gensim_segment_wiki",
                f"mediawiki_gensim_articles_streamed:{article_count}",
                f"external_mediawiki_gensim_interrupted:{type(exc).__name__}",
            ]
            return total_bytes, warnings
        try:
            out_path.unlink()
        except OSError:
            pass
        try:
            anchors_path.unlink()
        except OSError:
            pass
        return 0, [f"external_mediawiki_gensim_failed:{type(exc).__name__}"]
    if progress_callback is not None:
        progress_callback(safe_size(path), total_bytes)
    warnings = [
        "external_extractor:gensim_segment_wiki",
        f"mediawiki_gensim_articles_streamed:{article_count}",
    ]
    if article_count == 0:
        try:
            out_path.unlink()
        except OSError:
            pass
        try:
            anchors_path.unlink()
        except OSError:
            pass
        return 0, ["mediawiki_gensim_no_articles", *warnings]
    return total_bytes, warnings


def clean_anchor_path(clean_path: Path) -> Path:
    return clean_path.with_name(f"{clean_path.name}.anchors.jsonl")


def is_mediawiki_xml_source(path: Path) -> bool:
    try:
        head = read_bytes(path, max_bytes=128 * 1024)
    except OSError:
        return False
    lowered = head.lower()
    return b"<mediawiki" in lowered and b"<page>" in lowered


def iter_gensim_mediawiki_articles(path: Path, *, min_clean_chars: int, progress_callback=None):
    for clean, _anchors in iter_gensim_mediawiki_articles_with_anchors(
        path,
        min_clean_chars=min_clean_chars,
        progress_callback=progress_callback,
    ):
        yield clean


def iter_gensim_mediawiki_articles_with_anchors(path: Path, *, min_clean_chars: int, progress_callback=None):
    module = importlib.import_module("gensim.scripts.segment_wiki")
    extract_page_xmls = getattr(module, "extract_page_xmls")
    segment = getattr(module, "segment")
    ignored_namespaces = tuple(getattr(module, "IGNORED_NAMESPACES", ()))
    with path.open("rb") as source:
        for page_xml in extract_page_xmls(source):
            title, sections = segment(page_xml)
            if title and ignored_namespaces and any(str(title).startswith(f"{namespace}:") for namespace in ignored_namespaces):
                continue
            if not sections:
                continue
            section_bodies = [str(body or "") for _, body in sections]
            if section_bodies and section_bodies[0].lstrip().lower().startswith("#redirect"):
                continue
            clean, anchors = compose_gensim_mediawiki_article_payload(str(title or ""), sections)
            if len(clean) < min_clean_chars:
                continue
            if progress_callback is not None:
                try:
                    source_pos = source.tell()
                except OSError:
                    source_pos = 0
                progress_callback(source_pos, len(clean.encode("utf-8")) + 2)
            yield clean, anchors


def compose_gensim_mediawiki_article_text(title: str, sections) -> str:
    clean, _anchors = compose_gensim_mediawiki_article_payload(title, sections)
    return clean


def compose_gensim_mediawiki_article_payload(title: str, sections) -> tuple[str, list[dict]]:
    parts: list[str] = []
    anchors: list[dict] = []
    title = " ".join(str(title or "").split()).strip()
    if title:
        parts.append(title)
        anchors.append({"surface_kind": "title", "text": title})
    for heading, body in sections:
        heading_text = " ".join(str(heading or "").split()).strip()
        body_text = normalize_text(str(body or ""))
        if heading_text and heading_text.lower() != "introduction":
            parts.append(heading_text)
        if body_text:
            parts.append(body_text)
    text = normalize_text("\n\n".join(parts))
    text = strip_wiki_markup(text)
    clean = "\n".join(line for line in text.splitlines() if keep_wiki_payload_line(line))
    clean_anchors: list[dict] = []
    for anchor in anchors:
        anchor_text = clean_surface_anchor_text(str(anchor.get("text") or ""))
        if anchor_text:
            clean_anchors.append({"surface_kind": anchor.get("surface_kind") or "anchor", "text": anchor_text})
    return clean, clean_anchors


def clean_surface_anchor_text(text: str) -> str:
    clean = strip_wiki_markup(normalize_text(str(text or "")))
    clean = " ".join(clean.split()).strip()
    if len(clean.encode("utf-8", errors="ignore")) > 512:
        return ""
    return clean


def write_article_surface_anchors(
    handle,
    *,
    source_path: Path,
    clean_path: Path,
    article_text: str,
    article_start_offset: int,
    article_anchors: list[dict],
) -> int:
    count = 0
    search_start = 0
    for anchor in article_anchors:
        text = clean_surface_anchor_text(str(anchor.get("text") or ""))
        if not text:
            continue
        position = article_text.find(text, search_start)
        if position < 0:
            position = article_text.find(text)
        if position < 0:
            continue
        local_prefix = article_text[:position].encode("utf-8", errors="ignore")
        write_clean_anchor_record(
            handle,
            source_path=source_path,
            clean_path=clean_path,
            surface_kind=str(anchor.get("surface_kind") or "anchor"),
            text=text,
            content_offset=article_start_offset + len(local_prefix),
        )
        search_start = position + len(text)
        count += 1
    return count


def write_text_surface_anchors(*, source_path: Path, clean_path: Path, clean_text: str) -> int:
    anchors_path = clean_anchor_path(clean_path)
    count = 0
    byte_offset = 0
    first_payload_seen = False
    with anchors_path.open("w", encoding="utf-8") as handle:
        for raw_line in clean_text.splitlines(keepends=True):
            line_bytes = raw_line.encode("utf-8", errors="ignore")
            text = raw_line.strip()
            clean = clean_surface_anchor_text(text.lstrip("#").strip())
            surface_kind = ""
            if clean and not first_payload_seen:
                surface_kind = "title"
                first_payload_seen = True
            elif text.startswith("#") and clean:
                surface_kind = "heading"
            if surface_kind:
                line_prefix = raw_line[: len(raw_line) - len(raw_line.lstrip())].encode("utf-8", errors="ignore")
                write_clean_anchor_record(
                    handle,
                    source_path=source_path,
                    clean_path=clean_path,
                    surface_kind=surface_kind,
                    text=clean,
                    content_offset=byte_offset + len(line_prefix),
                )
                count += 1
            byte_offset += len(line_bytes)
    if count == 0:
        try:
            anchors_path.unlink()
        except OSError:
            pass
    return count


def write_clean_anchor_record(
    handle,
    *,
    source_path: Path,
    clean_path: Path,
    surface_kind: str,
    text: str,
    content_offset: int,
) -> None:
    payload = text.encode("utf-8", errors="ignore")
    record = {
        "schema": ANCHOR_SCHEMA,
        "source_path": str(source_path),
        "clean_path": str(clean_path),
        "surface_kind": safe_lineage_token(surface_kind, "anchor"),
        "content_offset": int(content_offset),
        "content_len": len(payload),
        "text": text,
        "candidate_field_only": True,
        "answer_permission_granted": False,
        "truth_asserted": False,
    }
    handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True))
    handle.write("\n")


def local_name(tag: str) -> str:
    text = str(tag or "")
    if "}" in text:
        return text.rsplit("}", 1)[-1]
    return text


def stream_mediawiki_text_payloads_to_file(
    path: Path,
    out_path: Path,
    min_clean_chars: int,
    progress_callback=None,
) -> tuple[int, list[str]]:
    total_bytes = 0
    payload_count = 0
    skipped_non_main_count = 0
    skipped_redirect_count = 0
    buffer = bytearray()
    page = bytearray()
    in_page = False
    tail_keep = len(b"</page>") - 1
    scratch_size = 1024 * 1024
    last_progress_at = 0.0
    last_progress_bytes = 0
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("rb") as source, out_path.open("w", encoding="utf-8", errors="strict") as output:
        while True:
            chunk = source.read(scratch_size)
            if not chunk:
                break
            if progress_callback is not None:
                source_pos = source.tell()
                now = time.perf_counter()
                if source_pos - last_progress_bytes >= 16 * 1024 * 1024 or now - last_progress_at >= 2.0:
                    progress_callback(source_pos, total_bytes)
                    last_progress_at = now
                    last_progress_bytes = source_pos
            buffer.extend(chunk)
            while True:
                if not in_page:
                    open_at = buffer.find(b"<page>")
                    if open_at < 0:
                        retain_bytearray_tail(buffer, len(b"<page>") - 1)
                        break
                    del buffer[: open_at + len(b"<page>")]
                    in_page = True
                else:
                    close_at = buffer.find(b"</page>")
                    if close_at >= 0:
                        page.extend(buffer[:close_at])
                        del buffer[: close_at + len(b"</page>")]
                        decision = mediawiki_page_payload_decision(bytes(page))
                        if decision == "redirect":
                            skipped_redirect_count += 1
                            page.clear()
                            in_page = False
                            continue
                        if decision != "main":
                            skipped_non_main_count += 1
                            page.clear()
                            in_page = False
                            continue
                        clean = normalize_text(strip_wiki_markup(mediawiki_text_payload_from_page(bytes(page))))
                        page.clear()
                        in_page = False
                        if len(clean) >= min_clean_chars:
                            output.write(clean)
                            output.write("\n\n")
                            total_bytes += len(clean.encode("utf-8")) + 2
                            payload_count += 1
                            if progress_callback is not None:
                                progress_callback(source.tell(), total_bytes)
                    else:
                        if len(buffer) > tail_keep:
                            page.extend(buffer[:-tail_keep])
                            del buffer[:-tail_keep]
                        break
        if in_page and page:
            page.extend(buffer)
            decision = mediawiki_page_payload_decision(bytes(page))
            if decision == "redirect":
                skipped_redirect_count += 1
            elif decision != "main":
                skipped_non_main_count += 1
            else:
                clean = normalize_text(strip_wiki_markup(mediawiki_text_payload_from_page(bytes(page))))
                if len(clean) >= min_clean_chars:
                    output.write(clean)
                    output.write("\n\n")
                    total_bytes += len(clean.encode("utf-8")) + 2
                    payload_count += 1
    if progress_callback is not None:
        progress_callback(safe_size(path), total_bytes)
    warnings = [
        f"mediawiki_main_namespace_text_payload_streamed:{payload_count}",
        f"mediawiki_non_main_pages_skipped:{skipped_non_main_count}",
        f"mediawiki_redirect_pages_skipped:{skipped_redirect_count}",
        "wiki_markup_stripped",
    ]
    if payload_count == 0:
        try:
            out_path.unlink()
        except OSError:
            pass
        return 0, ["mediawiki_stream_no_main_namespace_payloads", *warnings]
    return total_bytes, warnings


def mediawiki_page_payload_decision(page_payload: bytes) -> str:
    if not page_payload:
        return "empty"
    if b"<redirect" in page_payload[:4096]:
        return "redirect"
    ns = first_xml_text_bytes(page_payload, b"ns").strip()
    if ns and ns != b"0":
        return "non_main_namespace"
    title = html.unescape(first_xml_text_bytes(page_payload, b"title").decode("utf-8", errors="replace")).strip()
    if mediawiki_title_is_transport_page(title):
        return "transport_title"
    if not mediawiki_text_payload_from_page(page_payload).strip():
        return "empty_text"
    return "main"


def mediawiki_title_is_transport_page(title: str) -> bool:
    head = str(title or "").strip().split(":", 1)[0].lower()
    return head in {
        "talk",
        "user",
        "user talk",
        "wikipedia",
        "wikipedia talk",
        "file",
        "file talk",
        "mediawiki",
        "mediawiki talk",
        "template",
        "template talk",
        "help",
        "help talk",
        "category",
        "category talk",
        "portal",
        "portal talk",
        "module",
        "module talk",
        "special",
    }


def first_xml_text_bytes(payload: bytes, tag: bytes) -> bytes:
    open_tag = b"<" + tag + b">"
    close_tag = b"</" + tag + b">"
    start = payload.find(open_tag)
    if start < 0:
        return b""
    start += len(open_tag)
    end = payload.find(close_tag, start)
    if end < 0:
        return b""
    return payload[start:end]


def mediawiki_text_payload_from_page(page_payload: bytes) -> str:
    title = html.unescape(first_xml_text_bytes(page_payload, b"title").decode("utf-8", errors="replace")).strip()
    open_at = page_payload.find(b"<text")
    if open_at < 0:
        return title
    gt_at = page_payload.find(b">", open_at)
    if gt_at < 0:
        return title
    close_at = page_payload.find(b"</text>", gt_at + 1)
    if close_at < 0:
        return title
    raw = page_payload[gt_at + 1 : close_at]
    payload = html.unescape(raw.decode("utf-8", errors="replace")).strip()
    if title and payload:
        return f"{title}\n\n{payload}"
    return payload or title


def retain_bytearray_tail(buffer: bytearray, keep: int) -> None:
    if len(buffer) > keep:
        del buffer[: len(buffer) - keep]


def extract_mediawiki_text_payloads(path: Path, *, max_bytes_per_file: int) -> list[str]:
    raw = read_bytes(path, max_bytes=max_bytes_per_file)
    payloads = []
    cursor = 0
    while True:
        open_at = raw.find(b"<page>", cursor)
        if open_at < 0:
            break
        page_start = open_at + len(b"<page>")
        close_at = raw.find(b"</page>", page_start)
        if close_at < 0:
            break
        page_payload = raw[page_start:close_at]
        if mediawiki_page_payload_decision(page_payload) == "main":
            payload = mediawiki_text_payload_from_page(page_payload)
            if payload.strip():
                payloads.append(payload)
        cursor = close_at + len(b"</page>")
    return payloads


def strip_wiki_markup(value: str) -> str:
    text = html.unescape(str(value or ""))
    text = remove_balanced_markup(text, "{{", "}}")
    text = remove_balanced_markup(text, "{|", "|}")
    text = re.sub(r"\[\[(?:File|Image|Category):[^\]]+\]\]", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<ref\b[^>]*>.*?</ref>", " ", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"<ref\b[^/]*/>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<!--.*?-->", " ", text, flags=re.DOTALL)
    text = re.sub(r"\[\[(?:[^|\]]*\|)?([^\]]+)\]\]", r"\1", text)
    text = re.sub(r"\[https?://[^\s\]]+\s*([^\]]*)\]", r"\1", text)
    text = re.sub(r"https?://\S+", " ", text)
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"'{2,}", "", text)
    text = re.sub(r"(?m)^\s*(\|-|!|\|)\s*", " ", text)
    text = re.sub(r"(?m)^\s*==+\s*(.*?)\s*==+\s*$", r"\1", text)
    text = re.sub(r"__[^_\n]+__", " ", text)
    return "\n".join(line for line in text.splitlines() if keep_wiki_payload_line(line))


def keep_wiki_payload_line(value: str) -> bool:
    line = " ".join(str(value or "").split()).strip()
    if not line:
        return False
    lower = line.lower()
    if "category:" in lower or lower.startswith("category "):
        return False
    if "referencesexternallinks" in lower.replace(" ", ""):
        return False
    if "thumb|" in lower or lower.startswith(("image:", "file:", "[file:", "[image:")):
        return False
    if re.fullmatch(r"[*\s|,.;:(){}\[\]_\-–—]+", line):
        return False
    table_markers = (
        "class=",
        "bgcolor=",
        "align=",
        "rowspan=",
        "colspan=",
        "style=",
        "vcard",
        "fn org",
        "adr",
        "region",
    )
    marker_hits = sum(1 for marker in table_markers if marker in lower)
    table_pipe_count = line.count("||")
    separator_load = line.count("|") + table_pipe_count + line.count("=")
    if table_pipe_count >= 1:
        return False
    if marker_hits and separator_load >= 2:
        return False
    if marker_hits >= 2:
        return False
    star_count = line.count("*")
    wordish_count = sum(1 for part in re.split(r"\s+", line) if re.search(r"[A-Za-zА-Яа-яІіЇїЄєҐґ]", part))
    if star_count >= 8 and wordish_count <= 4:
        return False
    return True


def remove_balanced_markup(value: str, open_token: str, close_token: str) -> str:
    text = str(value or "")
    out = []
    index = 0
    depth = 0
    while index < len(text):
        if text.startswith(open_token, index):
            depth += 1
            index += len(open_token)
            if depth == 1:
                out.append(" ")
            continue
        if depth > 0 and text.startswith(close_token, index):
            depth -= 1
            index += len(close_token)
            if depth == 0:
                out.append(" ")
            continue
        if depth == 0:
            out.append(text[index])
        index += 1
    return "".join(out)


def markup_to_text(text: str) -> str:
    parser = ReadableMarkupParser()
    try:
        parser.feed(html.unescape(str(text or "")).replace("><", ">\n<"))
        return parser.text()
    except Exception:
        return normalize_text(re.sub(r"<[^>]+>", " ", html.unescape(str(text or ""))))


def read_text(path: Path, *, max_bytes_per_file: int) -> str:
    return read_bytes(path, max_bytes=max_bytes_per_file).decode("utf-8", errors="replace")


def read_bytes(path: Path, *, max_bytes: int) -> bytes:
    with path.open("rb") as handle:
        if max_bytes > 0:
            return handle.read(max_bytes)
        return handle.read()


def normalize_text(value: str) -> str:
    lines = []
    for line in html.unescape(str(value or "")).replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        clean = " ".join(line.split()).strip()
        if clean:
            lines.append(clean)
    return "\n\n".join(lines)


def parse_extensions(value: str) -> set[str]:
    result = set()
    for item in str(value or "").split(","):
        item = item.strip().lower()
        if not item:
            continue
        result.add(item if item.startswith(".") else f".{item}")
    return result or set(DEFAULT_EXTENSIONS)


def safe_relative(path: Path, root: Path) -> Path:
    try:
        return path.relative_to(root)
    except ValueError:
        return Path(path.name)


def safe_size(path: Path) -> int:
    try:
        return path.stat().st_size
    except OSError:
        return 0


def file_sha256(path: Path, *, max_bytes: int) -> str:
    digest = hashlib.sha256()
    remaining = max_bytes
    with path.open("rb") as handle:
        while True:
            if max_bytes > 0:
                if remaining <= 0:
                    break
                block = handle.read(min(1024 * 1024, remaining))
                remaining -= len(block)
            else:
                block = handle.read(1024 * 1024)
            if not block:
                break
            digest.update(block)
    return digest.hexdigest()


def sha256_hex(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


if __name__ == "__main__":
    main()
